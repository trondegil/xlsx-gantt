"""
Tests for features and fixes introduced in 0.2.0:

* ``GanttStyle.day_names`` / ``week_label_format`` locale customisation
* Deprecated ``annotation_a_*`` aliases mapping to ``annotation_r_*``
* ``patch_solid_databars(sheet=None)`` patching every worksheet
* The ``xlsx-gantt`` CLI (``xlsx_gantt.cli.main``)
"""
from __future__ import annotations

import io
import json
import zipfile
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import DataBarRule

from xlsx_gantt import GanttChart, GanttStyle
from xlsx_gantt._xlsx_patch import patch_solid_databars
from xlsx_gantt.cli import main as cli_main

START = datetime(2026, 1, 12)  # a Monday
END   = datetime(2026, 1, 18)  # the following Sunday

SECTIONS = [
    {
        "name": "Build",
        "tasks": [
            {
                "name": "Code",
                "time_estimate": 5,
                "progress": 50,
                "ranges": [{"start": START, "end": END, "color": "0070C0"}],
                "annotations": {"Alice": "R"},
            }
        ],
    }
]

DATE_COL0 = 5  # first date column (after Activity/Task/Est./Progress)


def _load(chart: GanttChart):
    return load_workbook(io.BytesIO(chart.generate_excel_bytes())).active


# ── Locale / labels ─────────────────────────────────────────────────────

def test_custom_day_names_and_week_label():
    style = GanttStyle(
        day_names=("Ma", "Ti", "On", "To", "Fr", "Lø", "Sø"),
        week_label_format="Uke {week} ({year})",
    )
    ws = _load(GanttChart(SECTIONS, START, END, style=style))
    assert ws.cell(row=2, column=DATE_COL0).value == "Ma"       # Monday
    assert ws.cell(row=2, column=DATE_COL0 + 6).value == "Sø"   # Sunday
    assert ws.cell(row=1, column=DATE_COL0).value == "Uke 3 (2026)"


def test_default_day_names_unchanged():
    ws = _load(GanttChart(SECTIONS, START, END))
    assert ws.cell(row=2, column=DATE_COL0).value == "Mon"
    assert ws.cell(row=1, column=DATE_COL0).value == "Week 3"


# ── Deprecated annotation_a_* aliases ───────────────────────────────────

def test_annotation_a_aliases_map_to_r_fields():
    style = GanttStyle(annotation_a_bg="C8FFCC", annotation_a_fg="112233")
    assert style.annotation_r_bg == "C8FFCC"
    assert style.annotation_r_fg == "112233"
    # And the chart still renders without error
    GanttChart(SECTIONS, START, END, resource_names=["Alice"],
               style=style).generate_excel_bytes()


# ── Multi-sheet DataBar patch ───────────────────────────────────────────

def _two_sheet_workbook_bytes() -> bytes:
    wb = Workbook()
    rule = DataBarRule(start_type="num", start_value=0,
                       end_type="num", end_value=100, color="7030A0")
    ws1 = wb.active
    ws1["D4"] = 50
    ws1.conditional_formatting.add("D4:D10", rule)
    ws2 = wb.create_sheet("Second")
    ws2["D4"] = 75
    rule2 = DataBarRule(start_type="num", start_value=0,
                        end_type="num", end_value=100, color="7030A0")
    ws2.conditional_formatting.add("D4:D10", rule2)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_patch_all_sheets():
    buf = io.BytesIO(_two_sheet_workbook_bytes())
    patch_solid_databars(buf, "D4:D10", sheet=None)
    with zipfile.ZipFile(buf) as z:
        for name in ("xl/worksheets/sheet1.xml", "xl/worksheets/sheet2.xml"):
            xml = z.read(name).decode("utf-8")
            assert 'gradient="0"' in xml, name


def test_patch_named_sheet_only():
    buf = io.BytesIO(_two_sheet_workbook_bytes())
    patch_solid_databars(buf, "D4:D10")  # default: sheet1 only
    with zipfile.ZipFile(buf) as z:
        assert 'gradient="0"' in z.read("xl/worksheets/sheet1.xml").decode()
        assert 'gradient="0"' not in z.read("xl/worksheets/sheet2.xml").decode()


# ── GanttChart.from_json ────────────────────────────────────────────────

def test_from_json_builds_chart(tmp_path):
    cfg = tmp_path / "chart.json"
    cfg.write_text(json.dumps(CLI_CONFIG))
    chart = GanttChart.from_json(str(cfg))
    assert chart.project_name == "CLI Demo"
    assert chart.resource_names == ["Alice"]
    assert len(chart.dates) == 7
    ws = load_workbook(io.BytesIO(chart.generate_excel_bytes())).active
    assert ws.cell(row=2, column=DATE_COL0).value == "Mon"


def test_from_json_invalid_input(tmp_path):
    bad_json = tmp_path / "bad.json"
    bad_json.write_text("{not json")
    missing_key = tmp_path / "missing.json"
    missing_key.write_text(json.dumps({"sections": []}))
    bad_date = tmp_path / "bad_date.json"
    bad_date.write_text(json.dumps(dict(CLI_CONFIG, start_date="12/01/2026")))
    for path in (bad_json, missing_key, bad_date):
        try:
            GanttChart.from_json(str(path))
        except ValueError:
            pass
        else:
            raise AssertionError(f"{path.name} should raise ValueError")


# ── CLI ─────────────────────────────────────────────────────────────────

CLI_CONFIG = {
    "project_name": "CLI Demo",
    "start_date": "2026-01-12",
    "end_date": "2026-01-18",
    "resource_names": ["Alice"],
    "sections": [
        {
            "name": "Build",
            "tasks": [
                {
                    "name": "Code",
                    "time_estimate": 5,
                    "progress": 50,
                    "ranges": [
                        {"start": "2026-01-12", "end": "2026-01-15",
                         "color": "0070C0"}
                    ],
                    "annotations": {"Alice": "R"},
                }
            ],
        }
    ],
}


def test_cli_generates_xlsx(tmp_path):
    cfg = tmp_path / "chart.json"
    out = tmp_path / "out.xlsx"
    cfg.write_text(json.dumps(CLI_CONFIG))
    assert cli_main([str(cfg), "-o", str(out), "--theme", "ocean"]) == 0
    assert out.exists()
    assert zipfile.is_zipfile(out)
    ws = load_workbook(out).active
    assert ws.cell(row=2, column=DATE_COL0).value == "Mon"


def test_cli_rejects_bad_json(tmp_path, capsys):
    cfg = tmp_path / "bad.json"
    cfg.write_text("{not json")
    assert cli_main([str(cfg), "-o", str(tmp_path / "o.xlsx")]) == 1
    assert "error" in capsys.readouterr().err


def test_cli_rejects_missing_keys(tmp_path, capsys):
    cfg = tmp_path / "missing.json"
    cfg.write_text(json.dumps({"sections": []}))
    assert cli_main([str(cfg), "-o", str(tmp_path / "o.xlsx")]) == 1
    assert "start_date" in capsys.readouterr().err


def test_cli_rejects_bad_date(tmp_path, capsys):
    bad = dict(CLI_CONFIG, start_date="12/01/2026")
    cfg = tmp_path / "bad_date.json"
    cfg.write_text(json.dumps(bad))
    assert cli_main([str(cfg), "-o", str(tmp_path / "o.xlsx")]) == 1
    assert "start_date" in capsys.readouterr().err


def test_cli_missing_file(tmp_path, capsys):
    assert cli_main([str(tmp_path / "nope.json")]) == 1
    assert "cannot read" in capsys.readouterr().err
