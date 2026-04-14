"""
xlsx_gantt/chart.py
~~~~~~~~~~~~~~~~~~~~
Core Gantt chart builder.

The public entry-points are :meth:`GanttChart.generate_excel` (saves to
disk) and :meth:`GanttChart.generate_excel_bytes` (returns raw bytes for
in-memory use).

Internal rendering is split across focused private helpers so that each
concern — header rows, data rows, total row, conditional formatting,
borders — lives in its own method rather than one monolithic function.
"""
from __future__ import annotations

import os
from collections import OrderedDict
from dataclasses import dataclass
from datetime import timedelta
from io import BytesIO
from typing import Callable

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU

from ._xlsx_patch import patch_solid_databars
from .models import DateRange, Section, Task
from .style import GanttStyle


# ══════════════════════════════════════════════════════════════════════
# Private rendering context
# ══════════════════════════════════════════════════════════════════════

@dataclass
class _RenderCtx:
    """Pre-computed rendering state shared across all ``_write_*`` helpers."""

    s: GanttStyle

    # ── Column indices ────────────────────────────────────────────────
    COL_ACT:   int
    COL_TASK:  int
    COL_EST:   int
    COL_PROG:  int
    DATE_COL0: int
    n_dates:   int
    n_res:     int
    RES_COL0:  int
    AS_COL:    int
    LAST_COL:  int

    # ── Borders ───────────────────────────────────────────────────────
    border:     Border
    hdr_border: Border

    # ── Alignments ───────────────────────────────────────────────────
    center:  Alignment
    left:    Alignment
    wrap:    Alignment
    wrap_l:  Alignment
    rotated: Alignment

    # ── Resolved colour strings ───────────────────────────────────────
    hdr_bg:        str
    logo_bg:       str
    week_bg:       str
    col_label_bg:  str
    day_name_bg:   str
    title_fg:      str
    week_label_fg: str
    day_name_fg:   str
    date_num_fg:   str
    col_label_fg:  str
    res_hdr_fg:    str
    as_legend_fg:  str

    # ── Pre-built fills ───────────────────────────────────────────────
    dark_fill:     PatternFill
    logo_fill:     PatternFill
    week_fill:     PatternFill
    col_lbl_fill:  PatternFill
    weekend_fill:  PatternFill
    day_name_fill: PatternFill
    info_bg_fill:  PatternFill | None

    # ── Font factory callables ────────────────────────────────────────
    hdr_font:  Callable[..., Font]
    body_font: Callable[..., Font]


# ══════════════════════════════════════════════════════════════════════
# GanttChart
# ══════════════════════════════════════════════════════════════════════

class GanttChart:
    """
    Generates a styled Gantt chart as an Excel (.xlsx) file.

    Parameters
    ----------
    sections : list[Section | dict]
        Each section groups related tasks under a common *Activity* label.
        Accepts :class:`~xlsx_gantt.Section` dataclass instances **or**
        plain dicts (the original dict API is fully preserved)::

            # dict form (original API)
            {
                'name': str,
                'tasks': [
                    {
                        'name': str,
                        'time_estimate': float | None,
                        'progress': float | None,   # 0–100
                        'ranges': [
                            {'start': datetime, 'end': datetime,
                             'color': 'RRGGBB'}
                        ],
                        'annotations': {resource_name: 'R' | 'S', ...}
                    }, ...
                ]
            }

            # dataclass form
            Section(name="Design", tasks=[
                Task(name="Research", time_estimate=3, progress=80,
                     ranges=[DateRange(start=..., end=..., color="0070C0")],
                     annotations={"Alice": "R"}),
            ])

        You may freely mix both forms in the same list.

    start_date     : datetime
    end_date       : datetime
    project_name   : str
    resource_names : list[str]   — annotation column headers
    style          : GanttStyle  — visual configuration (uses defaults if omitted)
    logo_path      : str | None  — path to a logo image (PNG recommended)
    """

    def __init__(
        self,
        sections:       list,
        start_date,
        end_date,
        project_name:   str        = "Project Name",
        resource_names: list[str]  = None,
        style:          GanttStyle = None,
        logo_path:      str | None = None,
    ):
        self.sections       = sections if sections is not None else []
        self.start_date     = start_date
        self.end_date       = end_date
        self.project_name   = project_name
        self.resource_names = resource_names or []
        self.style          = style or GanttStyle()
        self.logo_path      = logo_path
        # Guard against start_date > end_date → empty list instead of error
        self.dates = [
            start_date + timedelta(days=i)
            for i in range(max(0, (end_date - start_date).days + 1))
        ]

    # ══════════════════════════════════════════════════════════════════
    # Public API
    # ══════════════════════════════════════════════════════════════════

    def generate_excel(self, filename: str) -> None:
        """Build the Gantt chart and save it to *filename* (.xlsx)."""
        buf = self._render()
        with open(filename, "wb") as fh:
            fh.write(buf.getvalue())

    def generate_excel_bytes(self) -> bytes:
        """
        Build the Gantt chart and return the raw .xlsx bytes.

        Useful for in-memory pipelines — streaming from a web server,
        attaching to an e-mail, caching in memory — without writing to
        disk::

            chart = GanttChart(sections, ...)
            xlsx_bytes = chart.generate_excel_bytes()

            # Flask / Django example
            response = HttpResponse(
                xlsx_bytes,
                content_type="application/vnd.openxmlformats-"
                             "officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = (
                'attachment; filename="gantt.xlsx"'
            )
        """
        return self._render().getvalue()

    # ══════════════════════════════════════════════════════════════════
    # Core render pipeline
    # ══════════════════════════════════════════════════════════════════

    def _render(self) -> BytesIO:
        """
        Build the complete workbook and return a patched .xlsx as
        :class:`io.BytesIO` (position reset to 0).
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Gantt Chart"

        ctx             = self._build_ctx()
        data_end_row    = self._calc_data_end_row()

        self._setup_dimensions(ws, ctx)
        self._write_header_rows(ws, ctx)
        self._write_column_label_row(ws, ctx, data_end_row)
        self._prefill_data_area(ws, ctx, data_end_row)

        total_row, total_est, prog_vals = self._write_data_rows(ws, ctx)

        self._write_total_row(ws, ctx, total_row, total_est, prog_vals)
        prog_cf_range = self._add_conditional_formatting(ws, ctx, total_row)
        self._apply_borders(ws, ctx, total_row)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        if prog_cf_range:
            patch_solid_databars(buf, prog_cf_range)

        return buf

    # ══════════════════════════════════════════════════════════════════
    # Context builder
    # ══════════════════════════════════════════════════════════════════

    def _build_ctx(self) -> _RenderCtx:
        """Pre-compute all styling objects and return a :class:`_RenderCtx`."""
        s = self.style
        n_dates = len(self.dates)
        n_res   = len(self.resource_names)

        # ── Column layout ─────────────────────────────────────────────
        COL_ACT   = 1
        COL_TASK  = 2
        COL_EST   = 3
        COL_PROG  = 4
        DATE_COL0 = 5
        RES_COL0  = DATE_COL0 + n_dates
        AS_COL    = RES_COL0  + n_res
        LAST_COL  = AS_COL if n_res > 0 else (DATE_COL0 + n_dates - 1)

        # ── Borders ───────────────────────────────────────────────────
        _thin       = Side(style="thin", color=s.border_color)
        _border     = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
        _hdr_thin   = Side(style="thin", color=s.header_border_color)
        _hdr_border = Border(left=_hdr_thin, right=_hdr_thin,
                             top=_hdr_thin,  bottom=_hdr_thin)

        # ── Alignments ────────────────────────────────────────────────
        _center  = Alignment(horizontal="center", vertical="center")
        _left    = Alignment(horizontal="left",   vertical="center")
        _wrap    = Alignment(horizontal="center", vertical="center", wrap_text=True)
        _wrap_l  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        _rotated = Alignment(horizontal="center", vertical="bottom", text_rotation=90)

        # ── Resolved colour strings ───────────────────────────────────
        hdr_bg       = s.header_bg
        logo_bg      = s._r(s.logo_bg,       s.header_bg)
        week_bg      = s._r(s.week_label_bg, s.header_bg)
        col_label_bg = s._r(s.col_label_bg,  s.header_bg)
        day_name_bg  = s._r(s.day_name_bg,   s.header_bg)

        title_fg       = s._r(s.title_fg,        s.header_fg)
        week_label_fg  = s._r(s.week_label_fg,   s.header_fg)
        day_name_fg    = s._r(s.day_name_fg,     s.header_fg)
        date_num_fg    = s._r(s.date_number_fg,  s.header_fg)
        col_label_fg   = s._r(s.col_label_fg,    s.header_fg)
        res_hdr_fg     = s._r(s.resource_hdr_fg, s.header_fg)
        as_legend_fg   = s._r(s.as_legend_fg,    s.header_fg)

        # ── Fill factories ────────────────────────────────────────────
        def _solid(color: str) -> PatternFill:
            return PatternFill("solid", fgColor=color)

        dark_fill     = _solid(hdr_bg)
        logo_fill     = _solid(logo_bg)
        week_fill     = _solid(week_bg)
        col_lbl_fill  = _solid(col_label_bg)
        weekend_fill  = _solid(s.weekend_bg)
        day_name_fill = _solid(day_name_bg)
        info_bg_fill  = _solid(s.col_info_bg) if s.col_info_bg else None

        # ── Font factories ────────────────────────────────────────────
        def _hdr_font(fg: str | None = None, size: float | None = None,
                      bold: bool = True) -> Font:
            return Font(name=s.font_name, bold=bold,
                        color=fg or s.header_fg,
                        size=size or s.font_size_hdr)

        def _body_font(fg: str = "000000", bold: bool = False,
                       size: float | None = None) -> Font:
            return Font(name=s.font_name, bold=bold,
                        color=fg, size=size or s.font_size_body)

        return _RenderCtx(
            s=s,
            COL_ACT=COL_ACT, COL_TASK=COL_TASK,
            COL_EST=COL_EST, COL_PROG=COL_PROG,
            DATE_COL0=DATE_COL0, n_dates=n_dates, n_res=n_res,
            RES_COL0=RES_COL0, AS_COL=AS_COL, LAST_COL=LAST_COL,
            border=_border, hdr_border=_hdr_border,
            center=_center, left=_left, wrap=_wrap,
            wrap_l=_wrap_l, rotated=_rotated,
            hdr_bg=hdr_bg, logo_bg=logo_bg, week_bg=week_bg,
            col_label_bg=col_label_bg, day_name_bg=day_name_bg,
            title_fg=title_fg, week_label_fg=week_label_fg,
            day_name_fg=day_name_fg, date_num_fg=date_num_fg,
            col_label_fg=col_label_fg, res_hdr_fg=res_hdr_fg,
            as_legend_fg=as_legend_fg,
            dark_fill=dark_fill, logo_fill=logo_fill, week_fill=week_fill,
            col_lbl_fill=col_lbl_fill, weekend_fill=weekend_fill,
            day_name_fill=day_name_fill, info_bg_fill=info_bg_fill,
            hdr_font=_hdr_font, body_font=_body_font,
        )

    # ══════════════════════════════════════════════════════════════════
    # Setup: dimensions & header pre-fill
    # ══════════════════════════════════════════════════════════════════

    def _calc_data_end_row(self) -> int:
        """Count the first row *after* all data rows (used for auto-filter)."""
        total_tasks = 0
        for sec in self.sections:
            if isinstance(sec, Section):
                tasks = sec.tasks
            elif isinstance(sec, dict):
                tasks = sec.get("tasks", [])
                if not isinstance(tasks, list):
                    tasks = []
            else:
                continue
            total_tasks += len(tasks)
        return 3 + total_tasks + 1

    def _setup_dimensions(self, ws, ctx: _RenderCtx) -> None:
        """Set column widths, row heights, and pre-fill header rows with base colour."""
        s = ctx.s

        ws.column_dimensions[get_column_letter(ctx.COL_ACT)].width  = s.col_width_activity
        ws.column_dimensions[get_column_letter(ctx.COL_TASK)].width = s.col_width_task
        ws.column_dimensions[get_column_letter(ctx.COL_EST)].width  = s.col_width_estimate
        ws.column_dimensions[get_column_letter(ctx.COL_PROG)].width = s.col_width_progress
        for i in range(ctx.n_dates):
            ws.column_dimensions[get_column_letter(ctx.DATE_COL0 + i)].width = s.col_width_date
        for i in range(ctx.n_res):
            ws.column_dimensions[get_column_letter(ctx.RES_COL0 + i)].width = s.col_width_resource
        if ctx.n_res > 0:
            ws.column_dimensions[get_column_letter(ctx.AS_COL)].width = s.col_width_as_legend

        ws.row_dimensions[1].height = s.row_height_hdr1
        ws.row_dimensions[2].height = s.row_height_hdr2
        ws.row_dimensions[3].height = s.row_height_hdr3

        # Pre-fill all three header rows with the base dark colour
        for r in range(1, 4):
            for c in range(1, ctx.LAST_COL + 1):
                ws.cell(row=r, column=c).fill = ctx.dark_fill

    # ══════════════════════════════════════════════════════════════════
    # Header rows (1 & 2)
    # ══════════════════════════════════════════════════════════════════

    def _write_header_rows(self, ws, ctx: _RenderCtx) -> None:
        """Write rows 1–2: logo, title, resource headers, week labels, day names."""
        self._write_logo_area(ws, ctx)
        self._write_title(ws, ctx)
        self._write_resource_headers(ws, ctx)
        self._write_week_labels(ws, ctx)
        self._write_day_names(ws, ctx)

    def _write_logo_area(self, ws, ctx: _RenderCtx) -> None:
        """Merge A1:B2, fill with logo colour, and embed the logo image if present."""
        ws.merge_cells(start_row=1, start_column=ctx.COL_ACT,
                       end_row=2,   end_column=ctx.COL_TASK)
        cell = ws.cell(row=1, column=ctx.COL_ACT)
        cell.fill      = ctx.logo_fill
        cell.alignment = ctx.center

        if not (self.logo_path and os.path.isfile(self.logo_path)):
            return

        try:
            s         = ctx.s
            logo_img  = XLImage(self.logo_path)
            area_w_px = int((s.col_width_activity + s.col_width_task) * 7)
            area_h_px = int((s.row_height_hdr1 + s.row_height_hdr2) * 4 / 3)
            padding   = 4
            target_w  = max(1, area_w_px - padding * 2)
            target_h  = max(1, area_h_px - padding * 2)
            orig_w, orig_h = logo_img.width, logo_img.height
            if orig_w and orig_h:
                scale = (
                    min(target_w / orig_w, target_h / orig_h)
                    * max(0.0, min(1.0, s.logo_scale))
                )
                logo_img.width  = int(orig_w * scale)
                logo_img.height = int(orig_h * scale)
            marker = AnchorMarker(
                col=0, colOff=pixels_to_EMU(max(0, s.logo_offset_x)),
                row=0, rowOff=pixels_to_EMU(max(0, s.logo_offset_y)),
            )
            size = XDRPositiveSize2D(
                pixels_to_EMU(logo_img.width),
                pixels_to_EMU(logo_img.height),
            )
            logo_img.anchor = OneCellAnchor(_from=marker, ext=size)
            ws.add_image(logo_img)
        except Exception as err:
            print(f"Warning: could not insert logo '{self.logo_path}': {err}")

    def _write_title(self, ws, ctx: _RenderCtx) -> None:
        """Write the project title into merged cells C1:D2."""
        ws.merge_cells(start_row=1, start_column=ctx.COL_EST,
                       end_row=2,   end_column=ctx.COL_PROG)
        c = ws.cell(row=1, column=ctx.COL_EST)
        c.value     = self.project_name
        c.font      = ctx.hdr_font(fg=ctx.title_fg, size=ctx.s.font_size_hdr_lg)
        c.fill      = ctx.dark_fill
        c.alignment = ctx.wrap

    def _write_resource_headers(self, ws, ctx: _RenderCtx) -> None:
        """Write rotated resource-name headers and the A/S legend column."""
        for i, rname in enumerate(self.resource_names):
            col = ctx.RES_COL0 + i
            ws.merge_cells(start_row=1, start_column=col,
                           end_row=2,   end_column=col)
            c = ws.cell(row=1, column=col)
            c.value     = rname
            c.font      = ctx.hdr_font(fg=ctx.res_hdr_fg, size=8)
            c.fill      = ctx.dark_fill
            c.alignment = ctx.rotated

        if ctx.n_res > 0:
            ws.merge_cells(start_row=1, start_column=ctx.AS_COL,
                           end_row=2,   end_column=ctx.AS_COL)
            c = ws.cell(row=1, column=ctx.AS_COL)
            c.value     = "R: Responsible\nS: Support"
            c.font      = ctx.hdr_font(fg=ctx.as_legend_fg, size=8)
            c.fill      = ctx.dark_fill
            c.alignment = ctx.wrap_l

    def _write_week_labels(self, ws, ctx: _RenderCtx) -> None:
        """Merge and label week bands in row 1 across the date columns."""
        week_groups: OrderedDict = OrderedDict()
        for i, d in enumerate(self.dates):
            key = d.isocalendar()[:2]
            if key not in week_groups:
                week_groups[key] = [ctx.DATE_COL0 + i, ctx.DATE_COL0 + i, key[1]]
            else:
                week_groups[key][1] = ctx.DATE_COL0 + i

        for _key, (sc, ec, wnum) in week_groups.items():
            if sc != ec:
                ws.merge_cells(start_row=1, start_column=sc,
                               end_row=1,   end_column=ec)
            c = ws.cell(row=1, column=sc)
            c.value     = f"Week {wnum}"
            c.font      = ctx.hdr_font(fg=ctx.week_label_fg)
            c.fill      = ctx.week_fill
            c.alignment = ctx.center

    def _write_day_names(self, ws, ctx: _RenderCtx) -> None:
        """Write day-name abbreviations (Mon … Sun) in row 2."""
        _abbr = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        for i, d in enumerate(self.dates):
            c = ws.cell(row=2, column=ctx.DATE_COL0 + i)
            c.value     = _abbr[d.weekday()]
            c.font      = ctx.hdr_font(fg=ctx.day_name_fg,
                                       size=ctx.s.font_size_hdr_sm)
            c.fill      = ctx.day_name_fill
            c.alignment = ctx.center

    # ══════════════════════════════════════════════════════════════════
    # Column-label row (3)
    # ══════════════════════════════════════════════════════════════════

    def _write_column_label_row(
        self, ws, ctx: _RenderCtx, data_end_row: int
    ) -> None:
        """Write column headers, date numbers, auto-filter, and freeze panes."""
        s = ctx.s

        for col, txt, aln in [
            (ctx.COL_ACT,  "Activity",  ctx.left),
            (ctx.COL_TASK, "Task",      ctx.left),
            (ctx.COL_EST,  "Time Est.", ctx.wrap),
            (ctx.COL_PROG, "Progress",  ctx.center),
        ]:
            c = ws.cell(row=3, column=col)
            c.value     = txt
            c.font      = ctx.hdr_font(fg=ctx.col_label_fg)
            c.fill      = ctx.col_lbl_fill
            c.alignment = aln

        for i, d in enumerate(self.dates):
            c = ws.cell(row=3, column=ctx.DATE_COL0 + i)
            c.value     = f"{d.day}.{d.month}"
            c.font      = Font(name=s.font_name, color=ctx.date_num_fg,
                               size=s.font_size_hdr_sm)
            c.fill      = ctx.col_lbl_fill
            c.alignment = ctx.center

        for col in range(ctx.RES_COL0, ctx.LAST_COL + 1):
            ws.cell(row=3, column=col).fill = ctx.col_lbl_fill

        ws.auto_filter.ref = (
            f"{get_column_letter(ctx.COL_ACT)}3:"
            f"{get_column_letter(ctx.COL_PROG)}{data_end_row}"
        )
        ws.freeze_panes = ws.cell(row=4, column=ctx.DATE_COL0)

    # ══════════════════════════════════════════════════════════════════
    # Data area pre-fill
    # ══════════════════════════════════════════════════════════════════

    def _prefill_data_area(
        self, ws, ctx: _RenderCtx, data_end_row: int
    ) -> None:
        """
        Pre-fill the *Time Est.* / *Progress* columns and resource columns
        with their background colours before individual cells are written.
        """
        s = ctx.s
        res_bg_fill = PatternFill("solid", fgColor=s.progress_bar_color)
        for row in range(4, data_end_row + 1):
            if ctx.info_bg_fill:
                ws.cell(row=row, column=ctx.COL_EST).fill  = ctx.info_bg_fill
                ws.cell(row=row, column=ctx.COL_PROG).fill = ctx.info_bg_fill
            for col in range(ctx.RES_COL0, ctx.LAST_COL + 1):
                ws.cell(row=row, column=col).fill = res_bg_fill

    # ══════════════════════════════════════════════════════════════════
    # Data rows
    # ══════════════════════════════════════════════════════════════════

    def _write_data_rows(
        self, ws, ctx: _RenderCtx
    ) -> tuple[int, float, list[float]]:
        """
        Iterate over sections and tasks, writing each row to the sheet.

        Returns
        -------
        tuple[int, float, list[float]]
            ``(next_empty_row, total_time_estimate, progress_values)``
        """
        current_row  = 4
        data_row_idx = 0
        total_est    = 0.0
        prog_vals:      list[float]          = []
        section_ranges: list[tuple[int,int]] = []

        for section in self.sections:
            # Normalise to (name, tasks) for both dict and Section
            if isinstance(section, Section):
                sec_name = section.name
                tasks    = section.tasks
            elif isinstance(section, dict):
                sec_name = section.get("name", "")
                tasks    = section.get("tasks", [])
                if not isinstance(tasks, list):
                    tasks = []
            else:
                continue  # skip non-dict / non-Section entries

            sec_start_row = current_row

            for t_idx, task in enumerate(tasks):
                # Normalise task to a plain dict for the renderer
                if isinstance(task, Task):
                    task_data = task.to_dict()
                elif isinstance(task, dict):
                    task_data = task
                else:
                    continue

                ws.row_dimensions[current_row].height = ctx.s.row_height_data
                self._write_task_row(
                    ws, ctx, current_row,
                    sec_name, task_data,
                    is_section_header=(t_idx == 0),
                    data_row_idx=data_row_idx,
                )

                # Accumulate totals
                est = task_data.get("time_estimate")
                if est is not None:
                    try:
                        total_est += est
                    except TypeError:
                        pass

                prog = task_data.get("progress")
                if prog is not None:
                    try:
                        prog_vals.append(float(prog))
                    except (TypeError, ValueError):
                        pass

                current_row  += 1
                data_row_idx += 1

            if len(tasks) > 1:
                section_ranges.append((sec_start_row, current_row - 1))

        # Merge Activity (col A) cells vertically within each section
        for (sr, er) in section_ranges:
            ws.merge_cells(start_row=sr, start_column=ctx.COL_ACT,
                           end_row=er,   end_column=ctx.COL_ACT)
            mc = ws.cell(row=sr, column=ctx.COL_ACT)
            mc.font      = ctx.body_font(fg=ctx.s.section_name_fg, bold=True)
            mc.alignment = Alignment(horizontal="left", vertical="top",
                                     wrap_text=True)

        return current_row, total_est, prog_vals

    def _write_task_row(
        self,
        ws,
        ctx: _RenderCtx,
        row: int,
        sec_name: str,
        task: dict,
        is_section_header: bool,
        data_row_idx: int,
    ) -> None:
        """Write a single task row: background, cells, bars, annotations."""
        s = ctx.s

        # ── Optional row background ───────────────────────────────────
        row_fill: PatternFill | None = None
        if is_section_header and s.section_name_bg:
            row_fill = PatternFill("solid", fgColor=s.section_name_bg)
        elif data_row_idx % 2 == 0 and s.row_bg_even:
            row_fill = PatternFill("solid", fgColor=s.row_bg_even)
        elif data_row_idx % 2 == 1 and s.row_bg_odd:
            row_fill = PatternFill("solid", fgColor=s.row_bg_odd)

        if row_fill:
            for i in range(ctx.n_res + (1 if ctx.n_res > 0 else 0)):
                ws.cell(row=row, column=ctx.RES_COL0 + i).fill = row_fill

        # ── Activity (column A) ───────────────────────────────────────
        c = ws.cell(row=row, column=ctx.COL_ACT)
        c.value     = sec_name if is_section_header else ""
        c.font      = (ctx.body_font(fg=s.section_name_fg, bold=True)
                       if is_section_header
                       else ctx.body_font(fg=s.task_fg))
        c.alignment = ctx.left

        # ── Task name ─────────────────────────────────────────────────
        c = ws.cell(row=row, column=ctx.COL_TASK)
        c.value     = task.get("name", "")
        c.font      = ctx.body_font(fg=s.task_fg)
        c.alignment = ctx.left

        # ── Time Estimate ─────────────────────────────────────────────
        est = task.get("time_estimate")
        c = ws.cell(row=row, column=ctx.COL_EST)
        if est is not None:
            c.value = est
        c.font      = ctx.body_font(fg=s.estimate_fg)
        c.alignment = ctx.center

        # ── Progress ──────────────────────────────────────────────────
        prog = task.get("progress")
        c = ws.cell(row=row, column=ctx.COL_PROG)
        if prog is not None:
            c.value = prog
        c.font      = ctx.body_font(fg=s.progress_value_fg)
        c.alignment = ctx.center

        # ── Weekend fills ─────────────────────────────────────────────
        for i, d in enumerate(self.dates):
            if d.weekday() >= 5:
                ws.cell(row=row, column=ctx.DATE_COL0 + i).fill = ctx.weekend_fill

        # ── Gantt bars ────────────────────────────────────────────────
        raw_ranges = task.get("ranges") or []
        if not isinstance(raw_ranges, list):
            raw_ranges = []
        for rng in raw_ranges:
            # Guard: skip non-dict entries (None, strings, numbers, …)
            if not isinstance(rng, dict):
                continue
            # Guard: skip ranges missing either date key
            if "start" not in rng or "end" not in rng:
                continue
            color = rng.get("color") or s.bar_color
            try:
                bar_fill = PatternFill("solid", fgColor=color)
            except Exception:
                bar_fill = PatternFill("solid", fgColor=s.bar_color)
            si = (rng["start"] - self.start_date).days
            ei = (rng["end"]   - self.start_date).days
            for off in range(max(0, si), min(ei + 1, ctx.n_dates)):
                ws.cell(row=row, column=ctx.DATE_COL0 + off).fill = bar_fill

        # ── Annotation cells ──────────────────────────────────────────
        ann = task.get("annotations") or {}
        if not isinstance(ann, dict):
            ann = {}
        for i, rname in enumerate(self.resource_names):
            role = ann.get(rname, "")
            c = ws.cell(row=row, column=ctx.RES_COL0 + i)
            c.value     = role
            c.alignment = ctx.center
            if role == "R":
                c.font = ctx.body_font(fg=s.annotation_a_fg)
                if s.annotation_a_bg:
                    c.fill = PatternFill("solid", fgColor=s.annotation_a_bg)
            elif role == "S":
                c.font = ctx.body_font(fg=s.annotation_s_fg)
                if s.annotation_s_bg:
                    c.fill = PatternFill("solid", fgColor=s.annotation_s_bg)
            else:
                c.font = ctx.body_font(fg=s.annotation_default_fg)

    # ══════════════════════════════════════════════════════════════════
    # Total row
    # ══════════════════════════════════════════════════════════════════

    def _write_total_row(
        self,
        ws,
        ctx: _RenderCtx,
        total_row: int,
        total_est: float,
        prog_vals: list[float],
    ) -> None:
        """Write the *Total* summary row below all data rows."""
        s = ctx.s
        ws.row_dimensions[total_row].height = s.row_height_total

        _total_bg  = s._r(s.total_row_bg, s.header_bg)
        _total_fg  = s._r(s.total_row_fg, s.header_fg)
        total_fill = PatternFill("solid", fgColor=_total_bg)
        total_font = Font(name=s.font_name, bold=True,
                          color=_total_fg, size=s.font_size_total)

        for col in range(1, ctx.LAST_COL + 1):
            ws.cell(row=total_row, column=col).fill = total_fill

        c = ws.cell(row=total_row, column=ctx.COL_ACT)
        c.value     = "Total"
        c.font      = total_font
        c.alignment = ctx.left

        ws.cell(row=total_row, column=ctx.COL_TASK).font = total_font

        c = ws.cell(row=total_row, column=ctx.COL_EST)
        c.value     = total_est
        c.font      = total_font
        c.alignment = ctx.center

        if prog_vals:
            c = ws.cell(row=total_row, column=ctx.COL_PROG)
            c.value     = round(sum(prog_vals) / len(prog_vals), 1)
            c.font      = total_font
            c.alignment = ctx.center

    # ══════════════════════════════════════════════════════════════════
    # Conditional formatting
    # ══════════════════════════════════════════════════════════════════

    def _add_conditional_formatting(
        self, ws, ctx: _RenderCtx, total_row: int
    ) -> str | None:
        """
        Add a solid DataBar rule on the *Progress* column.

        Returns the sqref string so the caller can pass it to
        :func:`~xlsx_gantt._xlsx_patch.patch_solid_databars`, or
        ``None`` if there are no data rows.
        """
        if total_row <= 4:
            return None
        sqref = (
            f"{get_column_letter(ctx.COL_PROG)}4:"
            f"{get_column_letter(ctx.COL_PROG)}{total_row - 1}"
        )
        ws.conditional_formatting.add(
            sqref,
            DataBarRule(
                start_type="num", start_value=0,
                end_type="num",   end_value=100,
                color=ctx.s.progress_bar_color,
                showValue=True,
            ),
        )
        return sqref

    # ══════════════════════════════════════════════════════════════════
    # Borders
    # ══════════════════════════════════════════════════════════════════

    def _apply_borders(self, ws, ctx: _RenderCtx, total_row: int) -> None:
        """Apply thin borders to every cell in the used range."""
        for row in range(1, total_row + 1):
            border = ctx.hdr_border if (row <= 3 or row == total_row) else ctx.border
            for col in range(1, ctx.LAST_COL + 1):
                ws.cell(row=row, column=col).border = border
