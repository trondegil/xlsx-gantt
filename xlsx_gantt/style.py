"""
xlsx_gantt/style.py
~~~~~~~~~~~~~~~~~~~~
Visual configuration dataclass for :class:`~xlsx_gantt.GanttChart`.

All colour fields accept a 6-digit hex string without a leading ``#``
(e.g. ``"FFC000"``).  Fields typed ``str | None`` fall back to a
primary-palette value at render time — see the field docstrings.
"""
from __future__ import annotations

from dataclasses import dataclass

from openpyxl.styles import PatternFill


@dataclass
class GanttStyle:
    """
    Full visual configuration for a :class:`~xlsx_gantt.GanttChart`.

    Every colour field accepts a 6-digit hex string (no ``#``).
    Fields that accept ``None`` mean "no fill / inherit from the primary
    palette" — e.g. ``week_label_bg=None`` falls back to ``header_bg``.

    Quick-start example::

        from xlsx_gantt import GanttChart, GanttStyle

        style = GanttStyle(
            header_bg        = "2C3E50",   # rows 1-2 background (logo/title/week/day)
            bar_color        = "E74C3C",
            row_bg_even      = "F8F8FF",
            section_name_bg  = "D0D0FF",
            annotation_r_bg  = "C8FFCC",
            annotation_s_bg  = "FFE4B5",
            # ── Row 3 (column-label + date-number row) ──
            col_label_bg     = "FFC000",   # amber background for the label/date row
            col_label_fg     = "000000",   # text colour for Activity/Task/… labels
            date_number_fg   = "000000",   # text colour for date numbers (e.g. "23.3")
        )
        chart = GanttChart(sections, ..., style=style)
    """

    # ── Primary palette ───────────────────────────────────────────────
    # These act as the base/fallback for all header sub-colours.
    header_bg:          str = "7030A0"  # header background (dark purple)
    header_fg:          str = "FFFFFF"  # header text colour
    weekend_bg:         str = "EFEFEF"  # weekend column fill in data rows
    bar_color:          str = "9966CC"  # default Gantt bar colour
    progress_bar_color: str = "7030A0"  # progress data-bar colour
    border_color:       str = "CCCCCC"  # thin-border colour (data rows)
    header_border_color: str = "888888"  # thin-border colour for header rows 1-3

    # ── Header sub-colours ────────────────────────────────────────────
    # None → falls back to the primary palette value shown in the comment.
    logo_bg:         str | None = None      # logo placeholder bg      (→ header_bg)
    title_fg:        str | None = None      # project title text       (→ header_fg)
    week_label_bg:   str | None = None      # week-band cell bg        (→ header_bg)
    week_label_fg:   str | None = None      # "Week N" text            (→ header_fg)
    day_name_bg:     str | None = "7030A0"  # Mon/Tue/… row bg         (→ header_bg)
    day_name_fg:     str | None = None      # Mon/Tue/… text           (→ header_fg)
    date_number_fg:  str | None = "000000"  # d.m date text
    col_label_bg:    str | None = "FFC000"  # Activity/Task/… row bg
    col_label_fg:    str | None = "000000"  # Activity/Task/… text
    resource_hdr_fg: str | None = None      # resource col header text (→ header_fg)
    as_legend_fg:    str | None = None      # A/S legend text          (→ header_fg)

    # ── Data row colours ─────────────────────────────────────────────
    row_bg_odd:       str | None = None     # odd data-row bg  (None = white/default)
    row_bg_even:      str | None = None     # even data-row bg (None = white/default)
    section_name_bg:  str | None = None     # first-row-of-section bg (None = default)
    col_info_bg:      str | None = None     # Time Est. + Progress column bg (None = white)
    section_name_fg:  str        = "000000" # section name bold text
    task_fg:          str        = "000000" # task name text
    estimate_fg:      str        = "000000" # time-estimate value text
    progress_value_fg:str        = "000000" # progress value text

    # ── Total row ────────────────────────────────────────────────────
    total_row_bg: str | None = None     # total-row background (→ header_bg)
    total_row_fg: str | None = None     # total-row text colour  (→ header_fg)

    # ── Annotation (R/S role) cells ───────────────────────────────────
    annotation_r_fg:  str        = "000000" # "R: Responsible" text
    annotation_s_fg:  str        = "000000" # "S: Support" text
    annotation_r_bg:  str | None = None     # "R" cell background (None = default)
    annotation_s_bg:  str | None = None     # "S" cell background (None = default)
    annotation_default_fg: str   = "000000" # any other role text
    # Deprecated pre-0.2 aliases for the "R" role fields; if set, they
    # override annotation_r_fg / annotation_r_bg (see __post_init__).
    annotation_a_fg:  str | None = None
    annotation_a_bg:  str | None = None

    # ── Logo ──────────────────────────────────────────────────────────
    logo_scale:       float = 0.7   # fraction of the A1:B2 area (0–1)
    logo_offset_x:    int   = 4     # horizontal offset in pixels from cell A1
    logo_offset_y:    int   = 4     # vertical offset in pixels from cell A1

    # ── Locale / labels ───────────────────────────────────────────────
    day_names: tuple[str, ...] = (          # Mon-first day abbreviations (row 2)
        "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun",
    )
    week_label_format: str = "Week {week}"  # row-1 band label; {week} and {year} available

    # ── Font ──────────────────────────────────────────────────────────
    font_name:        str   = "Calibri"
    font_size_hdr:    float = 9     # column-label / week / day headers
    font_size_hdr_sm: float = 7     # day-name abbreviations & date numbers
    font_size_hdr_lg: float = 12    # project title
    font_size_body:   float = 9     # data cells
    font_size_total:  float = 10    # total row

    # ── Column widths ─────────────────────────────────────────────────
    col_width_activity:  float = 20.0   # Activity / section
    col_width_task:      float = 25.0   # Task name
    col_width_estimate:  float = 15.0   # Time Estimate
    col_width_progress:  float = 15.0   # Progress
    col_width_date:      float = 3.0    # Each day column
    col_width_resource:  float = 4.5    # Annotation / resource columns
    col_width_as_legend: float = 15.0   # A/S legend column

    # ── Row heights ───────────────────────────────────────────────────
    row_height_hdr1:  float = 28.0  # Row 1 – week labels
    row_height_hdr2:  float = 13.0  # Row 2 – day names
    row_height_hdr3:  float = 13.0  # Row 3 – column labels + date numbers
    row_height_data:  float = 14.0  # Data rows
    row_height_total: float = 16.0  # Total row

    def __post_init__(self) -> None:
        # Honour the deprecated annotation_a_* names when given.
        if self.annotation_a_fg is not None:
            self.annotation_r_fg = self.annotation_a_fg
        if self.annotation_a_bg is not None:
            self.annotation_r_bg = self.annotation_a_bg

    # ── Internal helpers ──────────────────────────────────────────────
    def _r(self, value: str | None, fallback: str) -> str:
        """Return *value* if not None, else *fallback*."""
        return value if value is not None else fallback

    def _fill(self, color: str | None) -> PatternFill | None:
        """Return a solid PatternFill, or None if color is None."""
        return PatternFill('solid', fgColor=color) if color else None
